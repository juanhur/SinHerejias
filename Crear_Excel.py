
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from tabulate import tabulate
from openpyxl import load_workbook

def leer_excel_como_texto(uploaded_file):
    """
    Lee un archivo Excel subido como UploadedFile en Streamlit y extrae la información 
    de todas las hojas como texto estructurado en formato tabular.
    
    Args:
        uploaded_file (UploadedFile): Archivo subido mediante Streamlit.
    
    Returns:
        str: Contenido del archivo Excel en formato de texto tabular.
    """
    try:
        # Leer todas las hojas del archivo Excel
        hojas = pd.read_excel(uploaded_file, sheet_name=None)  # Devuelve un diccionario {nombre_hoja: DataFrame}
        texto_resultado = ""
        
        for nombre_hoja, data in hojas.items():
            texto_resultado += f"\n--- Hoja: {nombre_hoja} ---\n"
            # Convertir el DataFrame en una tabla bien delimitada
            texto_resultado += tabulate(data, headers="keys", tablefmt="grid", showindex=False)
            texto_resultado += "\n\n"
        
        return texto_resultado

    except Exception as e:
        return f"Error al leer el archivo Excel: {e}"

def generar_excel(codigo_epica, epica, codigo_historia, descripcion_historia, criterios_aceptacion, prioridad, estimacion_esfuerzo, responsable,nombre_archivo):
    # Crear diccionario para la primera hoja
    max_len = max(len(codigo_epica), len(epica), len(codigo_historia), len(descripcion_historia), len(criterios_aceptacion),
                  len(prioridad), len(estimacion_esfuerzo), len(responsable))
    #max_len2 = max(len(codigo_historia), len(descripcion_historia), len(codigo_tarea), len(descripcion_tarea), len(esfuerzo_tarea))
    estado=["Nuevo"] *  max_len
    fecha_entrega = [np.nan] * max_len
    dependencias = [np.nan] * max_len
    notas = [np.nan] * max_len
    # Relacionar las historias de usuario con las épicas
    codigo_epica_expanded = []
    descripcion_epica_expanded = []

    for historia in codigo_historia:
        # Extraer el código de la épica desde el código de la historia (asume el formato "AN-2024-0001-HU1")
        epica_code = "-".join(historia.split('-')[:4])  # Extraer las primeras tres partes del código
        # Buscar el índice correspondiente en las épicas
        if epica_code in codigo_epica:
            index = codigo_epica.index(epica_code)
            codigo_epica_expanded.append(codigo_epica[index])
            descripcion_epica_expanded.append(epica[index])  # Agregar la descripción de la épica
        else:
            # Si no se encuentra la épica, agregar valores vacíos
            codigo_epica_expanded.append(None)
            descripcion_epica_expanded.append(None)

    data = {
        'Codigo Épica': codigo_epica_expanded + [np.nan] * (max_len - len( codigo_epica_expanded )),
        'Épica': descripcion_epica_expanded + [np.nan] * (max_len - len(descripcion_epica_expanded)),
        'Título o Código de la Historia': codigo_historia + [np.nan] * (max_len - len(codigo_historia)),
        'Descripción de la Historia': descripcion_historia + [np.nan] * (max_len - len(descripcion_historia)),
        'Criterios de Aceptación': criterios_aceptacion + [np.nan] * (max_len - len(criterios_aceptacion)),
        'Prioridad': prioridad + [np.nan] * (max_len - len(prioridad)),
        'Estimación de Esfuerzo': estimacion_esfuerzo + [np.nan] * (max_len - len(estimacion_esfuerzo)),
        'Responsable': responsable + [np.nan] * (max_len - len(responsable)),
        'Estado': estado + [np.nan] * (max_len - len(estado)),
        'Fecha Estimada de Entrega': fecha_entrega + [np.nan] * (max_len - len(fecha_entrega)),
        'Dependencias (Opcional)': dependencias + [np.nan] * (max_len - len(dependencias)),
        'Notas Adicionales (Opcional)': notas + [np.nan] * (max_len - len(notas))
    }

    # Relacionar las tareas con las historias de usuario
    #codigo_historia_expanded = []
    #descripcion_historia_expanded = []
    #for tarea in codigo_tarea:
        #historia_code = "-".join(tarea.split('-')[:5])
        #if historia_code in codigo_historia:
            #index = codigo_historia.index(historia_code)
            #codigo_historia_expanded.append(codigo_historia[index])
            #descripcion_historia_expanded.append(descripcion_historia[index])

    #responsable = [np.nan] *  max_len2
    #estado_inicial = ["To Do"] *  max_len2
    
    #data2 = {
        #'Título o Código de la Historia': codigo_historia_expanded+ [np.nan]* (max_len2 - len(codigo_historia_expanded)),
        #'Descripción de la Historia': descripcion_historia_expanded+ [np.nan]* (max_len2 - len(codigo_historia_expanded)),
        #'Código de tarea': codigo_tarea + [np.nan] * (max_len2 - len( codigo_tarea)),
        #'Descripción de Tarea': descripcion_tarea+ [np.nan] * (max_len2 - len(descripcion_tarea)),
        #'Esfuerzo Estimado Tarea':esfuerzo_tarea+ [np.nan] * (max_len2 - len(esfuerzo_tarea)),
        #'Responsable': responsable,
        #'Estado Inicial': estado_inicial
    #}

    # Crear DataFrames
    df = pd.DataFrame(data)
    print("df1 impreso")
    #print(data2)
    df2 = pd.DataFrame()
    print("df2 impreso")

    # Verificar si el archivo ya existe
    if not nombre_archivo.endswith(".xlsx"):
        nombre_archivo += ".xlsx"
    try:
        with pd.ExcelWriter(nombre_archivo, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, index=False, sheet_name="Historias de Usuario")
            print("df1 impreso")
            df2.to_excel(writer, index=False, sheet_name="Tareas Asociadas")
            print("df2 impreso")
    except FileNotFoundError:
        with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Historias de Usuario")
            print("df impreso")
            df2.to_excel(writer, index=False, sheet_name="Tareas Asociadas")
            print("df2 impreso")

    # Ajustar columnas para ambas hojas
    wb = load_workbook(nombre_archivo)
    for sheet_name in ["Historias de Usuario", "Tareas Asociadas"]:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
    wb.save(nombre_archivo)
    return nombre_archivo
    print("Se generó el archivo Excel con dos hojas.")



import pandas as pd
import numpy as np
from openpyxl import load_workbook

def generar_excel_desde_json(json_data, nombre_archivo):
    # Usamos los datos proporcionados en la variable original
    historias = json_data  # Ahora directamente es la lista de diccionarios

    # Crear listas para los datos
    codigo_epica = []
    epica = []
    codigo_historia = []
    descripcion_historia = []
    criterios_aceptacion = []
    prioridad = []
    estimacion_esfuerzo = []
    responsable = []
    estado = []
    fecha_entrega = []  # Este campo no está en los datos proporcionados, lo dejamos vacío por ahora
    dependencias = []  # Este campo tampoco está en los datos proporcionados
    notas = []  # Este campo tampoco está en los datos proporcionados
    if len(historias)==1:
        print("entro")
        # Suponiendo que 'historias' es un diccionario
        claves = list(historias.keys())  # Convertir las claves a una lista
        primera_clave = claves[0] if claves else None

        # Verificar si la primera clave es diferente de 'Codigo de usuario'
        if primera_clave and primera_clave != "Codigo Epica":
            historias=historias[primera_clave] 
    with open("hisotrias.txt", "w") as archivo:
        archivo.write(str(historias))  # Escribe el contenido de la variable en el archivo
    if historias!="Agent stopped due to iteration limit or time limit.":
        for historia in historias:
            print(historia)  # Mostrar cada historia para debug

            # Modificar las claves de acuerdo con la estructura de los diccionarios
            codigo_epica.append(historia.get("Codigo Epica", np.nan))
            epica.append(historia.get("Epica", np.nan))
            codigo_historia.append(historia.get("Codigo de la Historia", np.nan))
            descripcion_historia.append(historia.get("Descripcion de la Historia", np.nan))  # El nombre de la clave tiene un carácter especial
            criterios_aceptacion.append(historia.get("Criterios de Aceptacion", np.nan))  # Otro carácter especial
            prioridad.append(historia.get("Prioridad", np.nan))
            estimacion_esfuerzo.append(historia.get("Estimacion de Esfuerzo", np.nan))  # Carácter especial
            responsable.append(historia.get("Responsable", np.nan))
            estado.append(historia.get("Estado", "Nuevo"))
            fecha_entrega.append(np.nan)  # No hay campo en el JSON proporcionado
            dependencias.append(np.nan)  # No hay campo en el JSON proporcionado
            notas.append(np.nan)  # No hay campo en el JSON proporcionado

        # Crear DataFrame
        data = {
            'Código Épica': codigo_epica,
            'Épica': epica,
            'Título o Código de la Historia': codigo_historia,
            'Descripción de la Historia': descripcion_historia,
            'Criterios de Aceptación': criterios_aceptacion,
            'Prioridad': prioridad,
            'Estimación de Esfuerzo': estimacion_esfuerzo,
            'Responsable': responsable,
            'Estado': estado,
            'Fecha Estimada de Entrega': fecha_entrega,
            'Dependencias (Opcional)': dependencias,
            'Notas Adicionales (Opcional)': notas
        }

        df = pd.DataFrame(data)

        # Verificar si el archivo ya existe
        if not nombre_archivo.endswith(".xlsx"):
            nombre_archivo += ".xlsx"

        try:
            with pd.ExcelWriter(nombre_archivo, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(writer, index=False, sheet_name="Historias de Usuario")
        except FileNotFoundError:
            with pd.ExcelWriter(nombre_archivo, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Historias de Usuario")

        # Ajustar columnas
        wb = load_workbook(nombre_archivo)
        sheet_name = "Historias de Usuario"
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2
        wb.save(nombre_archivo)

        return nombre_archivo
    else:
        return historias

