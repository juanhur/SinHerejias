
import streamlit as st
from  MongoDB  import  insertar_matriz_producto_multi,obtener_proyectos_puntuacion_codigo, extraer_codigo_epicas_por_codigo_proyecto,consultar_matriz_producto_global,extraer_epicas_por_codigo_proyecto,consultar_matriz_exis_proyecto_nombre,actualizar_un_elemento_matriz_n,borrar_proyecto_matriz,obtener_proyectos,consultar_nombres_proyectos,obtener_proyectos_puntuacion, consultar_epicas_por_codigo_proyecto, consultar_epicas_exis_proyecto,obtener_orden_matriz,agregar_info_pro,insertar_matriz_producto, consultar_matriz_producto, consultar_matriz_exis_proyecto,actualizar_un_elemento_matriz
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from streamlit_sortables import sort_items
from collections import defaultdict

def obtener_valores_posiciones(hoja, fila_fija=True):
    valores_posiciones = {}
    rango = hoja.max_column if fila_fija else hoja.max_row
    
    for i in range(1, rango + 1):
        valor_celda = hoja.cell(row=1, column=i).value if fila_fija else hoja.cell(row=i, column=1).value
        if valor_celda:
            valores_posiciones.setdefault(valor_celda, []).append(hoja.cell(row=1, column=i).coordinate if fila_fija else hoja.cell(row=i, column=1).coordinate)
    print("ojo"+str(valores_posiciones))
    return valores_posiciones

def combinar_celdas(hoja, valores_posiciones):
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    relleno = PatternFill(start_color="F3E0DA", end_color="F3E0DA", fill_type="solid")
    
    for valor, celdas in valores_posiciones.items():
        if len(celdas) > 1:
            hoja.merge_cells(f"{celdas[0]}:{celdas[-1]}")
            celda_inicio = hoja[celdas[0]]
            celda_inicio.value = valor
            celda_inicio.alignment = Alignment(horizontal='center', vertical='center')
            celda_inicio.border = borde
            celda_inicio.fill = relleno
            celda_inicio.font = Font(bold=True)

def Generar_doc_matriz(usuario,df,columnas_orden=None):
    lista_orden,proyectos_ord=obtener_orden_matriz(st.session_state.area)
    # Ordenar el DataFrame por proyecto
    print(columnas_orden)
    if columnas_orden is not None:
        df.columns = columnas_orden
    df_filtrado = df.sort_values(by=["proyecto"])
    # Ordenar las columnas según el índice de las filas (asumiendo que las columnas coinciden con el índice)
    #df_filtrado = df_filtrado.reindex(columns=lista_orden)
    #df_filtrado = df_filtrado.reindex(index=lista_orden)

    proyectos_count = proyectos_ord.value_counts()
    proyectos_count_ordenado = proyectos_count.reindex(proyectos_ord.unique(), fill_value=0)
    proyectos_count=proyectos_count_ordenado
    # Elimina la columna 'proyecto'
    df.drop(columns=['proyecto'], inplace=True)
    df=df_filtrado.drop(columns=['proyecto'])#, inplace=True)

    salida = "./archivos/matriz_priorizacion_"+usuario+".xlsx"
    with pd.ExcelWriter(salida, engine='openpyxl') as writer:
        df.to_excel(writer, index=True, startrow=1, startcol=1, sheet_name='Sheet1')

    wb = load_workbook(salida)
    sheet = wb.active

    last_column, last_row = 3, 3
    for proyecto, count in proyectos_count.items():
        for col in range(count):
            sheet.cell(row=1, column=last_column + col, value=proyecto)
        for row in range(count):
            sheet.cell(row=last_row + row, column=1, value=proyecto)
        last_column += count
        last_row += count

    valores_fila = obtener_valores_posiciones(sheet, fila_fija=True)
    valores_columna = obtener_valores_posiciones(sheet, fila_fija=False)
    combinar_celdas(sheet, valores_fila)
    combinar_celdas(sheet, valores_columna)

    relleno_fila_columna = PatternFill(start_color="F3E0DA", end_color="F3E0DA", fill_type="solid")
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=2, column=col).fill = relleno_fila_columna
    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=2).fill = relleno_fila_columna

    relleno_diagonal = PatternFill(start_color="225469", end_color="225469", fill_type="solid")
    fuente_blanca = Font(color="FFFFFF", bold=True)
    for i in range(1, min(sheet.max_row, sheet.max_column) + 1):
        celda = sheet.cell(row=i, column=i)
        celda.fill = relleno_diagonal
        celda.font = fuente_blanca
        if i == 1:
            celda.value = "PROYECTO"
        if i == 2:
            celda.value = "EPICAS"
    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for fila in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for celda in fila:
            celda.border = borde_fino
            celda.alignment = Alignment(wrap_text=True)

    alineacion_centrada = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for fila in range(1, sheet.max_row + 1):
        sheet.cell(row=fila, column=1).alignment = alineacion_centrada
        sheet.cell(row=fila, column=2).alignment = alineacion_centrada
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=col).alignment = alineacion_centrada
        sheet.cell(row=2, column=col).alignment = alineacion_centrada

    ultima_columna, ultima_fila = sheet.max_column, sheet.max_row
    relleno_gris = PatternFill(start_color="5A5C5C", end_color="5A5C5C", fill_type="solid")
    relleno_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fuente_negra = Font(color="000000", bold=True)

    for fila in range(3, ultima_fila + 1):
        celda_suma = sheet.cell(row=fila, column=ultima_columna + 1)
        celda_suma.value = f"=SUM(C{fila}:{get_column_letter(ultima_columna)}{fila})"
        celda_suma.font, celda_suma.fill = fuente_blanca, relleno_gris
        celda_suma.alignment = alineacion_centrada

    celda_total_suma = sheet.cell(row=ultima_fila + 1, column=ultima_columna + 1)
    celda_total_suma.value, celda_total_suma.font, celda_total_suma.fill = f"=SUM({get_column_letter(ultima_columna + 1)}3:{get_column_letter(ultima_columna + 1)}{ultima_fila})", fuente_negra, relleno_amarillo
    celda_total_suma.alignment = alineacion_centrada

    for fila in range(3, ultima_fila + 1):
        celda_porcentaje = sheet.cell(row=fila, column=ultima_columna + 2)
        celda_porcentaje.value = f"={get_column_letter(ultima_columna + 1)}{fila}/{celda_total_suma.coordinate}*100"
        celda_porcentaje.font, celda_porcentaje.fill = fuente_blanca, relleno_gris
        celda_porcentaje.alignment = alineacion_centrada

    celda_total_porcentaje = sheet.cell(row=ultima_fila + 1, column=ultima_columna + 2)
    celda_total_porcentaje.value, celda_total_porcentaje.font, celda_total_porcentaje.fill = "100%", fuente_negra, relleno_amarillo
    celda_total_porcentaje.alignment = alineacion_centrada

    wb.save(salida)
    wb.close()
    return salida


def encontrar_prefijo_comun(lista):
    """Devuelve el mayor prefijo común de todos los strings en la lista."""
    if not lista:
        return ""
    prefijo = lista[0]
    for string in lista[1:]:
        while not string.startswith(prefijo):
            prefijo = prefijo[:-1]
            if not prefijo:
                return ""
    return prefijo

import pandas as pd
def asignar_valor(valor):
        if valor == 'Must Have':
            return 2
        elif valor == 'Should Have':
            return 1.5
        elif valor == 'Could Have':
            return 1
        elif valor == "Won’t Have":
            return 0
        else:
            return 0  # Para valores no definidos

def crear_matriz(uploaded_file):
    save_folder = "archivos"
    os.makedirs(save_folder, exist_ok=True)
    # Leer el archivo de Excel cargado desde Streamlit
    df = pd.read_excel(uploaded_file)
    # Extraer la primera columna (Proyectos) y la segunda (Épicas)
    proyectos_array = df.iloc[:, 0].dropna().unique().astype(str)
    epicas_agregar = df.iloc[:, 1].dropna().unique().astype(str)
     # Determinar el nombre común del proyecto
    nombre_proyecto = encontrar_prefijo_comun(list(proyectos_array)).removesuffix("-EP")
    df['Prioridad(Valor)'] = df.iloc[:,5].apply(asignar_valor)
    df["Prioridad_x_Esfuerzo"] = df["Estimación de Esfuerzo"] * df["Prioridad(Valor)"]
    df= df.iloc[:, [1,3,6,(int(df.shape[1])-2),(int(df.shape[1])-1)]]
    # Aplicar la función a la columna 5 y crear una nueva columna 'Valor Asignado'
    # impacto/esfuerzo  impacto(pr2)/ numero de horas+personas
# Agrupar por 'Épica' y sumar las columnas numéricas
    df_agrupado = df.groupby("Épica").sum(numeric_only=True)
    st.write("📊 *Impacto de cada Épica*")
    st.write("📌 **Tip:** En cada columna puedes hacer clic en los **tres puntitos** para **ordenar** los datos o **fijar la columna** según tu necesidad.")
    st.dataframe(df_agrupado)
    matriz_df,total_prioridad,df_Prioridad=Estimar_priorizacion_proyecto_relativa_peso(df_agrupado)
    #matriz_df=Estimar_priorizacion_proyecto(df_agrupado) matriz sin considerar la proporcion dle proyecto,solo relativa
    # Mostrar la matriz actualizada
    st.write("📊 Matriz de Priorización del Proyecto:"+" **"+ nombre_proyecto+"**")
    
    nombre = st.text_input("**Ingresa el nombre del proyecto:**")
    confirmar_guardado = st.checkbox("Estoy seguro de que deseo guardar esta matriz. Se almacenará en la matriz principal de priorización con todos los proyectos actuales.")
    #st.dataframe(matriz_df)
    # Crear el botón de guardado
    if st.button("Guardar Matriz de Priorización del Proyecto"):
        if confirmar_guardado:
            if nombre:
                nombre=nombre.replace("/", "_")
                if consultar_matriz_exis_proyecto(nombre_proyecto):
                    st.warning("No se puede agregar,la matriz ya contiene el proyecto: "+nombre_proyecto)
                    
                else:
                    exist_epica,epic=consultar_epicas_exis_proyecto(epicas_agregar)
                    #if exist_epica:
                        #st.warning("No se puede agregar este proyecto,  la epica: "+epic+" ya existe en otro.")
                    #else:
                    if consultar_matriz_exis_proyecto_nombre(nombre):
                            st.warning("No se puede agregar este proyecto,  el nombre del  ya existe")
                    else:
                        agregar_info_pro(proyectos_array,epicas_agregar,nombre,nombre_proyecto,total_prioridad,df_Prioridad)
                        insertar_matriz_producto(matriz_df,nombre_proyecto)
                        st.success("Matriz guardada exitosamente en la matriz principal de priorización.")
            else:
                st.warning("Obligatorio insertar el nombre del proyecto"+nombre_proyecto)
        else:
            st.warning("Por favor, confirma que deseas guardar la matriz marcando la casilla de verificación.")

def crear_matriz_global( ):
    result = consultar_matriz_producto(st.session_state.area)
    try:
        df = pd.DataFrame(result)
        if not df.empty:
            print("El DataFrame tiene datos.")
        else:
            df=None
    except:
        df=None
    if df is not None:
        proyectos_unicos = df["proyecto"].unique()
        proyectos_seleccionados = st.multiselect("Selecciona proyectos:", proyectos_unicos)
        epicas_proyecto_codigo=extraer_codigo_epicas_por_codigo_proyecto(proyectos_unicos)
        epicas_proyecto=extraer_epicas_por_codigo_proyecto(proyectos_unicos)
        # Aplicar filtro
        if proyectos_seleccionados:
            lista_orden,proyectos_ord=obtener_orden_matriz(st.session_state.area)
            df =consultar_matriz_producto_global(st.session_state.area,epicas_proyecto,proyectos_unicos,epicas_proyecto_codigo,proyectos_ord)
            df.set_index("id_epica", inplace=True)
            df_Proyectos_T,df_Proyecto=obtener_proyectos_puntuacion_codigo(st.session_state.area,epicas_proyecto_codigo,proyectos_ord)
            matriz_suma = df_Proyectos_T + df_Proyecto
            matriz_suma=calcular_Matriz_rango_valores(matriz_suma)
            #st.dataframe(df)
            for index, row in df.iterrows():
                index_name= epicas_proyecto_codigo[row[0]][index]
                #print(index_name)
                if index_name in matriz_suma.index:  # Verificar si el índice existe en matriz_suma
                    for col_name, valor in row.items():
                        if valor == 0:
                            df.at[index, col_name] = matriz_suma.loc[index_name, col_name]

            df_filtrado=df[df['proyecto'].isin(proyectos_seleccionados)]
            matrices_proyectos,mapeo_proyectos=separar_matriz_global_por_proyectos (df_filtrado)
            # Diccionario para almacenar las ediciones
            matrices_editadas = {}
            # Mostrar cada matriz y permitir la edición
            for nombre, df in matrices_proyectos.items():
                st.write(f"### Matriz para el proyecto: {nombre}")
                renombrar={v: k for k, v in epicas_proyecto_codigo[mapeo_proyectos[nombre]].items()}
                renombrar['proyecto'] = None
                # Editor de datos sin la columna "proyecto"
                edited_df = st.data_editor(
                    df,
                    column_config=renombrar,
                    num_rows="dynamic",
                    use_container_width=True
                )
                # Guardar en el diccionario con el nombre del proyecto
                matrices_editadas[nombre] = edited_df

            # Botón para guardar todas las matrices en MongoDB
            if st.button("Guardar cambios en la matriz global"):
                keys = list(matrices_editadas.keys())

                # Eliminar 'proyecto' de todos excepto el último
                for i, name in enumerate(keys):
                    if 'proyecto' in matrices_editadas[name].columns and i != len(keys) - 1:
                        matrices_editadas[name] = matrices_editadas[name].drop(columns=['proyecto'])

                # Concatenar los DataFrames
                edited_df = pd.concat(matrices_editadas.values(), axis=1)
                #st.dataframe(edited_df)
                #st.write(edited_df['proyecto'].unique())
                insertar_matriz_producto_multi(edited_df,edited_df['proyecto'].unique())
                st.success("Todas las matrices han sido guardadas correctamente.")
                #st.rerun()
        else:
            st.warning("Selecciona al menos un proyecto para ver la matriz")
        

        if st.button("Exportar a un Excel"):
                ruta=None
                lista_orden,proyectos_ord=obtener_orden_matriz(st.session_state.area)
                result = consultar_matriz_producto_global(st.session_state.area,epicas_proyecto,proyectos_unicos,epicas_proyecto_codigo,proyectos_ord)
                orden_columnas=[valor for proyecto in proyectos_ord.unique() for valor in epicas_proyecto_codigo.get(proyecto, {}).values()]
                orden_columnas.append('proyecto')
                print(epicas_proyecto_codigo)
                print(orden_columnas)
                df1 = pd.DataFrame(result).set_index('id_epica')
                df1 = df1.reindex(columns=orden_columnas)
                #convertir  el nombre dela columan a nombre epicas en lsita
                dic_inverso = {v: k for sub_dic in epicas_proyecto_codigo.values() for k, v in sub_dic.items()}
                orden_columnas = [dic_inverso.get(item, item) for item in df1.columns]
                #st.dataframe(df1)
                ruta=Generar_doc_matriz(st.session_state["name"], df1,orden_columnas)
                #st.dataframe(df1)
                if ruta:
                    # Leer el archivo en modo binario
                    with open(ruta, "rb") as archivo:
                        archivo_binario = archivo.read()
                    # Nombre del archivo para la descarga
                    archivo_nombre = os.path.basename(ruta)

                    # Botón de descarga
                    st.download_button(
                        label="📥 Haz clic aquí para descargar el archivo",
                        data=archivo_binario,
                        file_name=archivo_nombre,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # Eliminar el archivo después de la descarga
                    os.remove(ruta)
                    st.success("✅ El archivo ha sido descargado y eliminado correctamente.") 
    else:
        st.warning("No tiene actualmenten proyectos.")


def crear_matriz_global_su_orden( ):
    result = consultar_matriz_producto(st.session_state.area)
    try:
        df = pd.DataFrame(result)
        if not df.empty:
            print("El DataFrame tiene datos.")
        else:
            df=None
    except:
        df=None
    if df is not None:
        lista_orden,proyectos_ord=obtener_orden_matriz(st.session_state.area)
        proyectos_unicos = df["proyecto"].unique()
        epicas_proyecto_codigo=extraer_codigo_epicas_por_codigo_proyecto(proyectos_unicos)
        df_Proyectos_T,df_Proyecto=obtener_proyectos_puntuacion_codigo(st.session_state.area,epicas_proyecto_codigo,proyectos_ord)
        ######################cambio para manejar proyectos ocn misma epica
        # df_Proyectos_T,df_Proyecto=obtener_proyectos_puntuacion(st.session_state.area,lista_orden)
        matriz_suma = df_Proyectos_T + df_Proyecto
        proyectos_seleccionados=consultar_nombres_proyectos(list(proyectos_ord))
        Proyectocod_nom = [f"{k}/{v}" for k, v in proyectos_seleccionados.items()]
        #df["proyecto"]=proyectos_ord.values
        sorted_items = sort_items(Proyectocod_nom)
        ponderaciones = np.linspace(5, 0.5, len(sorted_items))
        # Número total de elementos en la lista
        imp_proyectos = {item.split('/')[0]: valor for item, valor in zip(sorted_items , ponderaciones)}
        print (imp_proyectos)
        matriz_suma["proyecto"]=proyectos_ord.values
        #st.dataframe(matriz_suma)
        columnas_numericas = matriz_suma.select_dtypes(include=['float64', 'int']).columns
        print(columnas_numericas)
        # Multiplicar las columnas numéricas por el valor correspondiente en el diccionario
        matriz_suma[columnas_numericas] =  matriz_suma.apply(lambda row: row[columnas_numericas] * imp_proyectos[row["proyecto"]], axis=1)
        matriz_suma.drop(columns=['proyecto'], inplace=True)
        df=calcular_Matriz_rango_valores(matriz_suma)
        for i in range(len(df)):
            df.iloc[i, i] = 1  # Asignamos 1 a cada elemento en la diagonal
        df["proyecto"]=proyectos_ord.values
        #st.dataframe(df)
        #convertir  el nombre dela columan a nombre epicas en lsita
        dic_inverso = {v: k for sub_dic in epicas_proyecto_codigo.values() for k, v in sub_dic.items()}
        orden_columnas = [dic_inverso.get(item, item) for item in df.columns]
        
        #ruta=Generar_doc_matriz(st.session_state["name"], df,orden_columnas)
        return df,orden_columnas
    else:
        st.warning("No tiene actualmente proyectos.")
    
    
def crear_matriz_globa_sug():
    result = consultar_matriz_producto(st.session_state.area)
    #st.dataframe(result)
    try:
        df = pd.DataFrame(result)
        if not df.empty:
            print("El DataFrame tiene datos.")
        else:
            df=None
    except:
        df=None
    if df is not None:
        lista_orden,proyectos_ord=obtener_orden_matriz(st.session_state.area)
        df_Proyectos_T,df_Proyecto=obtener_proyectos_puntuacion(st.session_state.area,lista_orden)
        matriz_suma = df_Proyectos_T + df_Proyecto
        df=calcular_Matriz_rango_valores(matriz_suma)
        df["proyecto"]=proyectos_ord.values
        return df
    else:
        st.warning("No tiene actualmente proyectos.")
        return None


def Estimar_priorizacion_proyecto(df):
    # Crear la matriz de priorización relativa en función de la última columna (Valor3)
    df['Ultima_Columna'] = df['Prioridad_x_Esfuerzo']
    # Asegurarse de que 'Epica' es el índice del DataFrame
    # Ordenamos las épicas de acuerdo a la última columna (de mayor a menor)
    df_sorted = df.sort_values('Ultima_Columna', ascending=False)

    # Inicializar la matriz de priorización
    n = len(df_sorted)
    priorization_matrix = np.ones((n, n))

    # Llenar la matriz con las relaciones de importancia
    for i in range(n):
        for j in range(n):
            if i != j:
                A = df_sorted.iloc[i]['Ultima_Columna']
                B = df_sorted.iloc[j]['Ultima_Columna']
                ratio = A / B
                
                if ratio >= 4:
                    priorization_matrix[i, j] = 4
                elif ratio >= 2:
                    priorization_matrix[i, j] = 2
                elif ratio >= 1:
                    priorization_matrix[i, j] = 1
                elif ratio >= 0.5:
                    priorization_matrix[i, j] = 1/2
                else:
                    priorization_matrix[i, j] = 1/4
    # Crear un DataFrame para la matriz de priorización
    priorization_df = pd.DataFrame(priorization_matrix, 
                                index=df_sorted.index, 
                                columns=df_sorted.index)

    return priorization_df
def Estimar_priorizacion_proyecto_relativa_peso(df):
    # Calcular el total de Prioridad_x_Esfuerzo
    total_prioridad = df['Prioridad_x_Esfuerzo'].sum()

    # Calcular el peso de cada épica en relación al total
    df['Peso_Proyecto'] = df['Prioridad_x_Esfuerzo'] / total_prioridad

    # Ordenar las épicas según la última columna (de mayor a menor)
    df_sorted = df.sort_values('Prioridad_x_Esfuerzo', ascending=False)

    # Inicializar la matriz de priorización
    n = len(df_sorted)
    priorization_matrix = np.ones((n, n))

    # Llenar la matriz con las relaciones de importancia considerando el peso del proyecto
    for i in range(n):
        for j in range(n):
            if i != j:
                priorization_matrix[i, j] = 0
    # Crear un DataFrame para la matriz de priorización
    priorization_df = pd.DataFrame(priorization_matrix, 
                                index=df_sorted.index, 
                                columns=df_sorted.index)

    # Mostrar la matriz de priorización
    print("Matriz de priorización:")
    print( df_sorted )
    return priorization_df,total_prioridad, df_sorted
def calcular_Matriz_rango_valores(matriz_suma):

    # Obtener todos los valores excepto la diagonal
    valores = matriz_suma.values.flatten()
    valores = valores[valores != 1]  # Excluir la diagonal porque será 1

    # Calcular los quintiles (percentiles 20, 40, 60, 80, 100)
    quintiles = np.percentile(valores, [20, 40, 60, 80, 100])

    # Definir los valores a asignar en cada rango
    valores_asignados = [0.25, 0.5, 1, 2, 4]

    # Función para mapear valores según los quintiles
    def mapear_valor(x):
        for i in range(5):
            if x <= quintiles[i]:
                return valores_asignados[i]
        return 4  # Si el valor es el máximo, asignar 4

    # Aplicar la transformación a toda la matriz
    matriz_rangos = matriz_suma.map(mapear_valor)

    # Mantener la diagonal en 1
    np.fill_diagonal(matriz_rangos.values, 1)

    return matriz_rangos

# Función para actualizar df con los valores editados
def actualizar_df(df, edited_df):
    try:
        # Asegurarse de que ambos DataFrames tengan los mismos índices (filas)
        # Filtramos df para que tenga solo las filas de edited_df
        df_filtrado = df[df.index.isin(edited_df.index)]

        # Actualizar solo las filas filtradas
        for idx in edited_df.index:
            # Compara y actualiza solo las filas con el mismo índice
            for col in edited_df.columns:
                if df.at[idx, col] != edited_df.at[idx, col]:  # Si el valor es diferente, actualizamos
                    df.at[idx, col] = edited_df.at[idx, col]

    except Exception as e:
        st.error(e)

    return df

def cargar_matriz_global(uploaded_file ):
    df = pd.read_excel(uploaded_file)
    df = df.iloc[:-1]
    df = df.iloc[:, :-2]
        # Establecer la primera fila como los nombres de las columnas
    df.iloc[0, 0] = 'proyecto'
    #df.columns = df.iloc[0]
    #df.columns = df.iloc[0]
    #df = df.drop(0).reset_index(drop=True)
    lista_columnas_proyectos =list(df.columns)
    proyectos_existentes_excel=list(set(lista_columnas_proyectos))
    while "PROYECTO" in proyectos_existentes_excel:
        proyectos_existentes_excel.remove("PROYECTO")
    epicas_proyecto_codigo=extraer_codigo_epicas_por_codigo_proyecto(proyectos_existentes_excel)
    for i in range(1, len(lista_columnas_proyectos)):
        if "Unnamed" in lista_columnas_proyectos[i]:
            lista_columnas_proyectos[i] = lista_columnas_proyectos[i-1]
    for i in range(len(df.columns)):
        # Cambiar el valor de la primera fila según una condición o lo que necesites
        if lista_columnas_proyectos[i]!="PROYECTO":
            df.iloc[0, i] = epicas_proyecto_codigo[lista_columnas_proyectos[i]][df.iloc[0, i]]

    df.columns = df.iloc[0]
    df = df.rename(columns={'EPICAS': 'epica'})
    # Eliminar la primera fila, ya que ahora es redundante
    df = df.drop(0).reset_index(drop=True)
    for index, row in df.iterrows():
        if isinstance(row['proyecto'], str):  # Cuando encuentres un nuevo valor (AN-xxxx-xxxx)
            current_value = row['proyecto']
        elif current_value is not None:  # Si ya hay un valor de proyecto actual
            df.at[index, 'proyecto'] = current_value  # Actualiza el valor en la columna
        # Agrupar por la columna 'proyecto'
    grouped = df.groupby('proyecto')

    # Crear un diccionario con DataFrames para cada proyecto
    project_dfs = {project: grouped.get_group(project) for project in grouped.groups}
    Proyectos_registrados=obtener_proyectos(st.session_state.area)
    Proyectos_registrados = list({item['Proyecto'] for item in Proyectos_registrados})
    proyectos_no_registrados = [project for project in project_dfs.keys() if project not in Proyectos_registrados]
    if proyectos_no_registrados :
        proyectos_no_registrados= ", ".join(proyectos_no_registrados)
        st.error("No se puede cargar el archivo ,actualmente no se existen los siguientes proyectos: "+"*"+proyectos_no_registrados+"*")
    else:
        for project, project_df in project_dfs.items():
            project_df= project_df.set_index('epica')
            #st.dataframe(project_df)
            #st.write( project)
            borrar_proyecto_matriz(project)
            project=[project]
            insertar_matriz_producto_multi(project_df,project)
            #insertar_matriz_producto(project_df,project)
        st.success("Matriz Actualizada ")
def separar_matriz_global_por_proyectos (df):
    proyectos = obtener_proyectos(st.session_state.area)  # Suponiendo que esta función obtiene los proyectos
    # Diccionario para agrupar por proyecto
    proyectos_agrupados = defaultdict(lambda: {"Nombre_proyecto": "", "Proyecto": "", "Epicas": []})
    # Agrupamos los proyectos según la nueva estructura
    for item in proyectos:
        clave_proyecto = item["Proyecto"]  # Ahora buscamos por "proyecto"
        proyectos_agrupados[clave_proyecto]["Nombre_proyecto"] = item["Nombre_proyecto"]
        proyectos_agrupados[clave_proyecto]["Proyecto"] = clave_proyecto
        proyectos_agrupados[clave_proyecto]["Epicas"].append(item["Código Proyecto"])

    # Convertir a lista
    proyectos = list(proyectos_agrupados.values())
    matrices_proyectos = {}
    mapeo_proyectos = {}
    
    for proyecto in proyectos:
        nombre_proyecto = proyecto["Nombre_proyecto"]
        epicas_proyecto = proyecto["Epicas"]
        mapeo_proyectos[nombre_proyecto] = proyecto["Proyecto"]

        # Filtrar columnas que corresponden al proyecto
        columnas_presentes = [str(epica) for epica in epicas_proyecto if str(epica) in df.columns]  # Cambiado a string
        columnas_presentes.append("proyecto")
        if columnas_presentes:
            # Filtramos y nos aseguramos de mantener el mismo índice
            df_proyecto = df[columnas_presentes].copy()
            df_proyecto = df_proyecto.reindex(df.index)  # Mantener el orden original de filas
            
            matrices_proyectos[nombre_proyecto] = df_proyecto

    return matrices_proyectos, mapeo_proyectos
