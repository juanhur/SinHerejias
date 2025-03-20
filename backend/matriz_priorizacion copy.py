
import streamlit as st
from  MongoDB  import consultar_epicas_por_codigo_proyecto, consultar_epicas_exis_proyecto,obtener_orden_matriz,agregar_info_pro,insertar_matriz_producto, consultar_matriz_producto, consultar_matriz_exis_proyecto,actualizar_un_elemento_matriz
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

def obtener_valores_posiciones(hoja, fila_fija=True):
    valores_posiciones = {}
    rango = hoja.max_column if fila_fija else hoja.max_row
    
    for i in range(1, rango + 1):
        valor_celda = hoja.cell(row=1, column=i).value if fila_fija else hoja.cell(row=i, column=1).value
        if valor_celda:
            valores_posiciones.setdefault(valor_celda, []).append(hoja.cell(row=1, column=i).coordinate if fila_fija else hoja.cell(row=i, column=1).coordinate)
    print("ojo"+str(valores_posiciones))
    return valores_posiciones

def combinar_celdas(hoja, valores_posiciones):
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    relleno = PatternFill(start_color="F3E0DA", end_color="F3E0DA", fill_type="solid")
    
    for valor, celdas in valores_posiciones.items():
        if len(celdas) > 1:
            hoja.merge_cells(f"{celdas[0]}:{celdas[-1]}")
            celda_inicio = hoja[celdas[0]]
            celda_inicio.value = valor
            celda_inicio.alignment = Alignment(horizontal='center', vertical='center')
            celda_inicio.border = borde
            celda_inicio.fill = relleno
            celda_inicio.font = Font(bold=True)

def Generar_doc_matriz(usuario,df):
    lista_orden,proyectos_ord=obtener_orden_matriz(st.session_state.area)
    # Ordenar el DataFrame por proyecto
    df_filtrado = df.sort_values(by=["proyecto"])
    # Ordenar las columnas según el índice de las filas (asumiendo que las columnas coinciden con el índice)
    df_filtrado = df_filtrado.reindex(columns=lista_orden)
    df_filtrado = df_filtrado.reindex(index=lista_orden)

    proyectos_count = proyectos_ord.value_counts()
    proyectos_count_ordenado = proyectos_count.reindex(proyectos_ord.unique(), fill_value=0)
    proyectos_count=proyectos_count_ordenado
    # Elimina la columna 'proyecto'
    df.drop(columns=['proyecto'], inplace=True)
    df=df_filtrado

    salida = "./archivos/matriz_priorizacion_"+usuario+".xlsx"
    with pd.ExcelWriter(salida, engine='openpyxl') as writer:
        df.to_excel(writer, index=True, startrow=1, startcol=1, sheet_name='Sheet1')

    wb = load_workbook(salida)
    sheet = wb.active

    last_column, last_row = 3, 3
    for proyecto, count in proyectos_count.items():
        for col in range(count):
            sheet.cell(row=1, column=last_column + col, value=proyecto)
        for row in range(count):
            sheet.cell(row=last_row + row, column=1, value=proyecto)
        last_column += count
        last_row += count

    valores_fila = obtener_valores_posiciones(sheet, fila_fija=True)
    valores_columna = obtener_valores_posiciones(sheet, fila_fija=False)
    combinar_celdas(sheet, valores_fila)
    combinar_celdas(sheet, valores_columna)

    relleno_fila_columna = PatternFill(start_color="F3E0DA", end_color="F3E0DA", fill_type="solid")
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=2, column=col).fill = relleno_fila_columna
    for row in range(1, sheet.max_row + 1):
        sheet.cell(row=row, column=2).fill = relleno_fila_columna

    relleno_diagonal = PatternFill(start_color="225469", end_color="225469", fill_type="solid")
    fuente_blanca = Font(color="FFFFFF", bold=True)
    for i in range(1, min(sheet.max_row, sheet.max_column) + 1):
        celda = sheet.cell(row=i, column=i)
        celda.fill = relleno_diagonal
        celda.font = fuente_blanca
        if i == 1:
            celda.value = "PROYECTO"

    borde_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for fila in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for celda in fila:
            celda.border = borde_fino
            celda.alignment = Alignment(wrap_text=True)

    alineacion_centrada = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for fila in range(1, sheet.max_row + 1):
        sheet.cell(row=fila, column=1).alignment = alineacion_centrada
        sheet.cell(row=fila, column=2).alignment = alineacion_centrada
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row=1, column=col).alignment = alineacion_centrada
        sheet.cell(row=2, column=col).alignment = alineacion_centrada

    ultima_columna, ultima_fila = sheet.max_column, sheet.max_row
    relleno_gris = PatternFill(start_color="5A5C5C", end_color="5A5C5C", fill_type="solid")
    relleno_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fuente_negra = Font(color="000000", bold=True)

    for fila in range(3, ultima_fila + 1):
        celda_suma = sheet.cell(row=fila, column=ultima_columna + 1)
        celda_suma.value = f"=SUM(C{fila}:{get_column_letter(ultima_columna)}{fila})"
        celda_suma.font, celda_suma.fill = fuente_blanca, relleno_gris
        celda_suma.alignment = alineacion_centrada

    celda_total_suma = sheet.cell(row=ultima_fila + 1, column=ultima_columna + 1)
    celda_total_suma.value, celda_total_suma.font, celda_total_suma.fill = f"=SUM({get_column_letter(ultima_columna + 1)}3:{get_column_letter(ultima_columna + 1)}{ultima_fila})", fuente_negra, relleno_amarillo
    celda_total_suma.alignment = alineacion_centrada

    for fila in range(3, ultima_fila + 1):
        celda_porcentaje = sheet.cell(row=fila, column=ultima_columna + 2)
        celda_porcentaje.value = f"={get_column_letter(ultima_columna + 1)}{fila}/{celda_total_suma.coordinate}*100"
        celda_porcentaje.font, celda_porcentaje.fill = fuente_blanca, relleno_gris
        celda_porcentaje.alignment = alineacion_centrada

    celda_total_porcentaje = sheet.cell(row=ultima_fila + 1, column=ultima_columna + 2)
    celda_total_porcentaje.value, celda_total_porcentaje.font, celda_total_porcentaje.fill = "100%", fuente_negra, relleno_amarillo
    celda_total_porcentaje.alignment = alineacion_centrada

    wb.save(salida)
    wb.close()
    return salida

def encontrar_prefijo_comun(lista):
    """Devuelve el mayor prefijo común de todos los strings en la lista."""
    if not lista:
        return ""
    prefijo = lista[0]
    for string in lista[1:]:
        while not string.startswith(prefijo):
            prefijo = prefijo[:-1]
            if not prefijo:
                return ""
    return prefijo

import pandas as pd
def asignar_valor(valor):
        if valor == 'Must Have':
            return 2
        elif valor == 'Should Have':
            return 1.5
        elif valor == 'Could Have':
            return 1
        elif valor == "Won’t Have":
            return 0
        else:
            return 0  # Para valores no definidos

def crear_matriz(uploaded_file):
    save_folder = "archivos"
    os.makedirs(save_folder, exist_ok=True)
    # Leer el archivo de Excel cargado desde Streamlit
    df = pd.read_excel(uploaded_file)
    # Extraer la primera columna (Proyectos) y la segunda (Épicas)
    proyectos_array = df.iloc[:, 0].dropna().unique().astype(str)
    epicas_agregar = df.iloc[:, 1].dropna().unique().astype(str)
     # Determinar el nombre común del proyecto
    nombre_proyecto = encontrar_prefijo_comun(list(proyectos_array)).removesuffix("-EP")
    df['Prioridad(Valor)'] = df.iloc[:,5].apply(asignar_valor)
    df["Prioridad_x_Esfuerzo"] = df["Estimación de Esfuerzo"] * df["Prioridad(Valor)"]
    df= df.iloc[:, [1,3,6,(int(df.shape[1])-2),(int(df.shape[1])-1)]]
    # Aplicar la función a la columna 5 y crear una nueva columna 'Valor Asignado'
    # impacto/esfuerzo  impacto(pr2)/ numero de horas+personas
# Agrupar por 'Épica' y sumar las columnas numéricas
    df_agrupado = df.groupby("Épica").sum(numeric_only=True)
    st.write("📊 *Impacto de cada Épica*")
    st.write("📌 **Tip:** En cada columna puedes hacer clic en los **tres puntitos** para **ordenar** los datos o **fijar la columna** según tu necesidad.")
    st.dataframe(df_agrupado)
    matriz_df,total_prioridad=Estimar_priorizacion_proyecto_relativa_peso(df_agrupado)
    #matriz_df=Estimar_priorizacion_proyecto(df_agrupado) matriz sin considerar la proporcion dle proyecto,solo relativa
    # Mostrar la matriz actualizada
    st.write("📊 Matriz de Priorización del Proyecto:"+" **"+ nombre_proyecto+"**")

    nombre = st.text_input("**Ingresa el nombre del proyecto:**")
    st.write(
        """📌 **Tip:** 1) para editar una celda solo es necesario hacer doble click en ella."""
        )
    edited_df = st.data_editor(
        matriz_df,
        num_rows="dynamic",
        use_container_width=True
    )
            # Crear un checkbox para la confirmación
    confirmar_guardado = st.checkbox("Estoy seguro de que deseo guardar esta matriz. Se almacenará en la matriz principal de priorización con todos los proyectos actuales.")

    # Crear el botón de guardado
    if st.button("Guardar Matriz de Priorización del Proyecto"):
        if confirmar_guardado:
            if nombre:
                nombre=nombre.replace("/", "_")
                agregar_info_pro(proyectos_array,epicas_agregar,nombre,nombre_proyecto,total_prioridad)
                if consultar_matriz_exis_proyecto(nombre_proyecto):
                    st.warning("No se puede agregar,la matriz ya contiene el proyecto: "+nombre_proyecto)
                    
                else:
                    exist_epica,epic=consultar_epicas_exis_proyecto(epicas_agregar)
                    if exist_epica:
                        st.warning("No se puede agregar este proyecto,  la epica: "+epic+" ya existe en otro.")
                    else:
                        agregar_info_pro(proyectos_array,epicas_agregar,nombre,nombre_proyecto,total_prioridad)
                        insertar_matriz_producto(matriz_df,nombre_proyecto)
                        st.success("Matriz guardada exitosamente en la matriz principal de priorización.")
            else:
                st.warning("Obligatorio insertar el nombre del proyecto"+nombre_proyecto)
        else:
            st.warning("Por favor, confirma que deseas guardar la matriz marcando la casilla de verificación.")




def crear_matriz_global( ):
    result = consultar_matriz_producto(st.session_state.area)
    try:
        df = pd.DataFrame(result).drop(columns=['_id']).set_index('epica')
    except:
        df=None
    if df is not None:
        proyectos_unicos = df["proyecto"].unique()
        proyectos_seleccionados = st.multiselect("Selecciona proyectos:", proyectos_unicos)
        # Aplicar filtro
        if proyectos_seleccionados:
            lista_orden,proyectos_ord=obtener_orden_matriz(st.session_state.area)
            epicas_proyecto_sel=consultar_epicas_por_codigo_proyecto(proyectos_seleccionados)
            df_filtrado = df[df["proyecto"].isin(proyectos_seleccionados)]
            df_filtrado = df_filtrado.sort_values(by=["proyecto"])
           # Ordenar las columnas según la lista_orden
            df_filtrado = df_filtrado.reindex(columns=lista_orden)
            lista_filtrada = [elemento for elemento in lista_orden if elemento in epicas_proyecto_sel]
            # Ordenar las filas según el índice de lista_orden (suponiendo que lista_orden también contiene los índices de las filas)
            df_filtrado = df_filtrado.reindex(index= lista_filtrada)
            
            # Mostrar el DataFrame sin la columna "proyecto"
            edited_df = st.data_editor(
                df_filtrado.loc[:, ~df_filtrado.columns.isin(["proyecto"])],
                num_rows="dynamic",
                use_container_width=True
            )

                # Botón para guardar cambios en MongoDB
            if st.button("Guardar cambios en la matriz global"):
                for index, row in edited_df.iterrows():
                    # Buscar si la épica (índice) existe en la colección
                    filtro = {"epica": index}
                    actualizacion = {"$set": row.to_dict()}  # Convierte la fila editada a diccionario
                    # Actualizar en MongoDB
                    actualizar_un_elemento_matriz(filtro,actualizacion)

                st.success("✅ Datos actualizados correctamente en MongoDB.")
        
        else:
            st.warning("Selecciona al menos un proyecto para ver la matriz")
        

        if st.button("Exportar a Excel"):
                ruta=None
                ruta=Generar_doc_matriz(st.session_state["name"],df)
                if ruta:
                    # Leer el archivo en modo binario
                    with open(ruta, "rb") as archivo:
                        archivo_binario = archivo.read()
                    # Nombre del archivo para la descarga
                    archivo_nombre = os.path.basename(ruta)

                    # Botón de descarga
                    st.download_button(
                        label="📥 Haz clic aquí para descargar el archivo",
                        data=archivo_binario,
                        file_name=archivo_nombre,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    # Eliminar el archivo después de la descarga
                    os.remove(ruta)
                    st.success("✅ El archivo ha sido descargado y eliminado correctamente.") 
    else:
        st.warning("No tiene actualmenten proyectos.")


def Estimar_priorizacion_proyecto(df):
    # Crear la matriz de priorización relativa en función de la última columna (Valor3)
    df['Ultima_Columna'] = df['Prioridad_x_Esfuerzo']
    # Asegurarse de que 'Epica' es el índice del DataFrame
    # Ordenamos las épicas de acuerdo a la última columna (de mayor a menor)
    df_sorted = df.sort_values('Ultima_Columna', ascending=False)

    # Inicializar la matriz de priorización
    n = len(df_sorted)
    priorization_matrix = np.ones((n, n))

    # Llenar la matriz con las relaciones de importancia
    for i in range(n):
        for j in range(n):
            if i != j:
                A = df_sorted.iloc[i]['Ultima_Columna']
                B = df_sorted.iloc[j]['Ultima_Columna']
                ratio = A / B
                
                if ratio >= 4:
                    priorization_matrix[i, j] = 4
                elif ratio >= 2:
                    priorization_matrix[i, j] = 2
                elif ratio >= 1:
                    priorization_matrix[i, j] = 1
                elif ratio >= 0.5:
                    priorization_matrix[i, j] = 1/2
                else:
                    priorization_matrix[i, j] = 1/4
    # Crear un DataFrame para la matriz de priorización
    priorization_df = pd.DataFrame(priorization_matrix, 
                                index=df_sorted.index, 
                                columns=df_sorted.index)

    return priorization_df
def Estimar_priorizacion_proyecto_relativa_peso(df):
    # Calcular el total de Prioridad_x_Esfuerzo
    total_prioridad = df['Prioridad_x_Esfuerzo'].sum()

    # Calcular el peso de cada épica en relación al total
    df['Peso_Proyecto'] = df['Prioridad_x_Esfuerzo'] / total_prioridad

    # Ordenar las épicas según la última columna (de mayor a menor)
    df_sorted = df.sort_values('Prioridad_x_Esfuerzo', ascending=False)

    # Inicializar la matriz de priorización
    n = len(df_sorted)
    priorization_matrix = np.ones((n, n))

    # Llenar la matriz con las relaciones de importancia considerando el peso del proyecto
    for i in range(n):
        for j in range(n):
            if i != j:
                A = df_sorted.iloc[i]['Prioridad_x_Esfuerzo']
                B = df_sorted.iloc[j]['Prioridad_x_Esfuerzo']
                ratio = A / B
                weight_A = df_sorted.iloc[i]['Peso_Proyecto']
                weight_B = df_sorted.iloc[j]['Peso_Proyecto']

                # Calcular la relación de importancia ponderada por el peso del proyecto
                weighted_ratio = ratio * (weight_A / weight_B)
                priorization_matrix[i, j] =weighted_ratio
                # Asignar los valores de priorización según la relación ponderada
                #if weighted_ratio >= 6:
                    #priorization_matrix[i, j] = 4
                #elif weighted_ratio >= 3:
                    #priorization_matrix[i, j] = 2
                #elif weighted_ratio >= 1:
                    #priorization_matrix[i, j] = 1
                #elif weighted_ratio >= 0.5:
                    #priorization_matrix[i, j] = 1/2
                #else:
                    #priorization_matrix[i, j] = 1/4

    # Crear un DataFrame para la matriz de priorización
    priorization_df = pd.DataFrame(priorization_matrix, 
                                index=df_sorted.index, 
                                columns=df_sorted.index)

    # Mostrar la matriz de priorización
    print("Matriz de priorización:")
    print( df_sorted )
    return priorization_df,total_prioridad
